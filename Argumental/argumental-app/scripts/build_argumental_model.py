"""Build the Argumental investor financial model spreadsheet.

Reads the default scenario from ../src/lib/financialModelInputs.json
(the same file that powers /deck and /model on the website) and writes
a 5-tab xlsx to ../public/argumental-financial-model.xlsx.

Run:
    python3 argumental-app/scripts/build_argumental_model.py

Tabs:
    1. Read Me            — orientation
    2. Assumptions        — every input the model uses
    3. Bouts & Audience   — viewer ramp, replay reach, votes
    4. P&L (3-Year)       — revenue, COGS, gross margin, opex, EBITDA
    5. Unit Economics     — per-bout contribution margin
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = Path(__file__).resolve().parent
APP_ROOT = HERE.parent
INPUTS_PATH = APP_ROOT / "src" / "lib" / "financialModelInputs.json"
OUT_PATH = APP_ROOT / "public" / "argumental-financial-model.xlsx"

# ── Read the default scenario ────────────────────────────────────────────
with INPUTS_PATH.open("r") as f:
    DATA = json.load(f)

DEFAULT_KEY = DATA["_meta"]["defaultScenario"]
SCENARIO = DATA["scenarios"][DEFAULT_KEY]
SCENARIO_LABEL = SCENARIO["label"]

A = SCENARIO["audience"]
SP = SCENARIO["sponsorship"]
PM = SCENARIO["premium"]
TK = SCENARIO["ticketing"]
MR = SCENARIO["merch"]
LC = SCENARIO["licensing"]
AS_ = SCENARIO["adShare"]
V = SCENARIO["variableCosts"]
FX = SCENARIO["fixedCosts"]
B = SCENARIO["audienceBehavior"]
H = V["honorarium"]
T = SCENARIO["targets"]

# ── Style ────────────────────────────────────────────────────────────────
RED = "EB2C35"
BLUE = "1165C6"
INK = "18181B"
GREY = "71717A"
PALE = "F4F4F5"
PALE_BLUE = "EAF1FB"

THIN = Side(border_style="thin", color="E4E4E7")

H1 = Font(name="Helvetica Neue", size=22, bold=True, color=INK)
H2 = Font(name="Helvetica Neue", size=12, bold=True, color="FFFFFF")
KICKER = Font(name="Helvetica Neue", size=9, bold=True, color=GREY)
LABEL = Font(name="Helvetica Neue", size=11, bold=True, color=INK)
ROW_FONT = Font(name="Helvetica Neue", size=11, color=INK)
NOTE = Font(name="Helvetica Neue", size=9, italic=True, color=GREY)
INPUT_FONT = Font(name="Helvetica Neue", size=11, bold=True, color=BLUE)
CALC_FONT = Font(name="Helvetica Neue", size=11, color=INK)
TOTAL_FONT = Font(name="Helvetica Neue", size=11, bold=True, color=INK)

FILL_INPUT = PatternFill("solid", fgColor=PALE_BLUE)
FILL_TOTAL = PatternFill("solid", fgColor=PALE)

USD = '"$"#,##0;[Red]("$"#,##0)'
PCT = "0.0%"


def section_header(ws, row, text, color=INK, span_to=5):
    ws.cell(row=row, column=1, value=text).font = H2
    for c in range(1, span_to + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(
            horizontal="left" if c == 1 else "center",
            vertical="center",
            indent=1 if c == 1 else 0,
        )
    ws.row_dimensions[row].height = 22


def kicker(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = KICKER


def title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1
    ws.row_dimensions[row].height = 32


def label_cell(ws, row, text, *, total=False):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TOTAL_FONT if total else LABEL
    cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")


def input_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = INPUT_FONT
    c.fill = FILL_INPUT
    c.alignment = Alignment(horizontal="right")
    if fmt:
        c.number_format = fmt
    c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def calc_cell(ws, row, col, value, fmt=None, *, total=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = TOTAL_FONT if total else CALC_FONT
    c.alignment = Alignment(horizontal="right")
    if fmt:
        c.number_format = fmt
    if total:
        c.fill = FILL_TOTAL


def note(ws, row, text):
    ws.cell(row=row, column=5, value=text).font = NOTE


def header_row(ws, row, cols, color=INK):
    for c, val in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = H2
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(
            horizontal="left" if c == 1 else "right",
            vertical="center",
            indent=1 if c == 1 else 0,
        )
    ws.row_dimensions[row].height = 22


def write_input_row(ws, row, label, values, fmt, note_text):
    label_cell(ws, row, label)
    input_cell(ws, row, 2, values[0], fmt)
    input_cell(ws, row, 3, values[1], fmt)
    input_cell(ws, row, 4, values[2], fmt)
    note(ws, row, note_text)


# ── Workbook ─────────────────────────────────────────────────────────────
wb = Workbook()

# ── Tab 1: Read Me ───────────────────────────────────────────────────────
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110

kicker(ws, 1, f"ARGUMENTAL · INVESTOR MODEL · 3-YEAR · {SCENARIO_LABEL.upper()} CASE")
title(ws, 2, "How this model works")

readme = [
    "",
    f"This workbook renders the {SCENARIO_LABEL} scenario, the default case",
    "shown on /deck and /model. Inputs come from",
    "src/lib/financialModelInputs.json — the same file that drives the",
    "TypeScript views.  Edit the JSON, re-run this script, and all three",
    "(deck · dashboard · spreadsheet) stay in lockstep.",
    "",
    "Tabs",
    "  1. Read Me — this page.",
    "  2. Assumptions — every input the model uses, in one place.",
    "     Pale-blue cells = inputs; everything else is a formula.",
    "  3. Bouts & Audience — viewer ramp, replay reach, votes.",
    "  4. P&L (3-Year) — full revenue stack, variable costs, gross margin,",
    "     fixed opex, EBITDA, year-over-year + 3-Yr total.",
    "  5. Unit Economics — per-bout contribution margin, three audience scales.",
    "",
    "Year 1 stated goal",
    f"  · {T['liveViewersEOY1']:,} live viewers per bout by end of Year 1.",
    "",
    "Reach assumption — the 41× multiplier",
    f"  · {1 + B['replayMultiplier'][0]}× total impressions per bout (live + post-live replay).",
    f"  · {AS_['offPlatformPctOfReplay']:.0%} of replay viewing happens off-platform on FB / TikTok / YouTube.",
    "    Mux delivers only the on-platform portion; off-platform impressions",
    "    earn ad share from the host platforms.",
    "",
    f"What the {SCENARIO_LABEL.lower()} case has built in",
    "  · Voting revenue at $5 / vote, $10 / week cap, gated by live reach.",
    "  · Sponsorship — title slots ramping from Y1 to full cadence in Y3.",
    "  · Premium subscription — $7 / month archive + AMA tier.",
    "  · Live championship ticketing — physical events Y2 and Y3.",
    "  · Merch + international format licensing.",
    "  · YouTube ad share at $1.50 CPM on off-platform replay views.",
    "  · Off-platform replay distribution (saves Mux delivery cost).",
    "  · Honoraria: Y1 flat $10K, Y2/Y3 revenue share (5% / 3% of bout purse).",
    "  · Charity payout 18% of voting revenue, accounted as COGS.",
    "  · Mux enterprise rate by Y3 ($0.025 / viewer-hr).",
    "  · Lean headcount: 4 → 7 → 10 FTE.",
    "",
    "What this model deliberately omits",
    "  · Working-capital and payment-processor float.",
    "  · Tax — model is pre-tax EBITDA only.",
    "",
    f"For the conservative floor case (voting only, full headcount, no",
    "off-platform replay), see /model?scenario=conservative on the website.",
]

for i, line in enumerate(readme, start=3):
    cell = ws.cell(row=i, column=1, value=line)
    if (
        line.endswith("Conventions")
        or line.endswith("Tabs")
        or line.startswith("What")
        or line.startswith("Year 1")
        or line.startswith("Reach")
    ):
        cell.font = LABEL
    else:
        cell.font = ROW_FONT


# ── Tab 2: Assumptions ───────────────────────────────────────────────────
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 50
for col in "BCDE":
    ws.column_dimensions[col].width = 18

kicker(ws, 1, f"INPUTS — {SCENARIO_LABEL.upper()} SCENARIO")
title(ws, 2, "Assumptions")
header_row(ws, 4, ["Driver", "2026 (Y1)", "2027 (Y2)", "2028 (Y3)", "Notes"])

# Audience
section_header(ws, 6, "AUDIENCE & ENGAGEMENT")
write_input_row(ws, 7, "Bouts per year",
                A["boutsPerYear"], "#,##0", "1 / Sunday")
write_input_row(ws, 8, "Avg concurrent live viewers / bout",
                A["liveViewers"], "#,##0",
                f"Y1 EOY peak target: {T['liveViewersEOY1']:,}")
write_input_row(ws, 9, "% of viewers who vote",
                A["voterConversion"], "0.0%", "Single $5 vote")
write_input_row(ws, 10, "Avg vote price (USD)",
                A["votePrice"], '"$"#,##0', "Fixed")

# Sponsorship
section_header(ws, 12, "SPONSORSHIP")
write_input_row(ws, 13, "Sponsor revenue / bout (USD)",
                SP["sponsorPerBout"], '"$"#,##0', "Title slot · 41× reach")
write_input_row(ws, 14, "Sponsored bouts (% of cadence)",
                SP["sponsoredPct"], "0%", "Brand fit")

# Variable costs
section_header(ws, 16, "VARIABLE COSTS  ·  PER BOUT")
write_input_row(ws, 17, "Mux delivery $/viewer-hour",
                V["muxDeliveryRate"], '"$"#,##0.000',
                "Y3 = enterprise rate")
write_input_row(ws, 18, "Mux ingest $/bout (2 streams)",
                V["muxIngestPerBout"], '"$"#,##0', "24m × 2 streams")
write_input_row(ws, 19, "Stripe rate (% of vote rev)",
                V["stripePct"], "0.0%", "+ $0.30/vote")
write_input_row(ws, 20, "Stripe per-vote fee",
                V["stripePerVote"], '"$"0.00', "Flat")
write_input_row(ws, 21, "Charity payout (% of vote rev)",
                V["charityPct"], "0%", "18% donor model")

# Honorarium row — Y1 flat input, Y2/Y3 formulas
label_cell(ws, 22, "Debater honorarium / bout")
input_cell(ws, 22, 2, H["y1FlatUSD"], '"$"#,##0')
calc_cell(ws, 22, 3, f"=C8*C9*C10*{H['y2PctOfBoutPurse']}", '"$"#,##0')
calc_cell(ws, 22, 4, f"=D8*D9*D10*{H['y3PctOfBoutPurse']}", '"$"#,##0')
note(ws, 22,
     f"Y1 flat ${H['y1FlatUSD']:,} · Y2 {H['y2PctOfBoutPurse']:.0%} · "
     f"Y3 {H['y3PctOfBoutPurse']:.0%} of purse")

write_input_row(ws, 23, "Production / bout",
                V["productionPerBout"], '"$"#,##0', "Studio + crew")

# Fixed costs
section_header(ws, 25, "FIXED COSTS  ·  PER YEAR")
write_input_row(ws, 26, "Headcount (FTE)",
                FX["headcount"], "#,##0", "Avg loaded ~$185K")
write_input_row(ws, 27, "Loaded cost per FTE",
                FX["loadedCostPerFTE"], '"$"#,##0', "Salary + benefits")
write_input_row(ws, 28, "Paid search & acquisition",
                FX["paidSearch"], '"$"#,##0', "$10K/mo Y1")
write_input_row(ws, 29, "Brand & creator marketing",
                FX["brandMarketing"], '"$"#,##0', "Sponsored clips")
write_input_row(ws, 30, "G&A, tooling, infra",
                FX["gAndA"], '"$"#,##0', "Office, software")
write_input_row(ws, 31, "Legal, accounting, insurance",
                FX["legalAccounting"], '"$"#,##0', "Outside counsel")

# Audience behavior
section_header(ws, 33, "AUDIENCE BEHAVIOR")
write_input_row(ws, 34, "Avg minutes watched / live viewer",
                B["avgLiveMins"], "#,##0", "Drives Mux live-delivery cost")
write_input_row(ws, 35, "Post-live replay multiplier",
                B["replayMultiplier"], '#,##0"x"', "Replay views = live × this")
write_input_row(ws, 36, "Avg minutes watched / replay viewer",
                B["avgReplayMins"], "#,##0", "Highlights & clips")

# Platform revenue & distribution
section_header(ws, 38, "PLATFORM REVENUE & DISTRIBUTION")
write_input_row(ws, 39, "Premium subscribers (year-end)",
                PM["subscribersByYear"], "#,##0", "Archive + AMAs")
label_cell(ws, 40, "Premium price / mo")
for c in (2, 3, 4):
    input_cell(ws, 40, c, PM["monthlyPrice"], '"$"#,##0')
note(ws, 40, "Flat across years")

write_input_row(ws, 41, "Live championship events / yr",
                TK["eventsByYear"], "#,##0", "Physical venue")
label_cell(ws, 42, "Seats per event")
for c in (2, 3, 4):
    input_cell(ws, 42, c, TK["seatsPerEvent"], "#,##0")
label_cell(ws, 43, "Avg ticket price")
for c in (2, 3, 4):
    input_cell(ws, 43, c, TK["avgTicketPrice"], '"$"#,##0')
write_input_row(ws, 44, "Merch annual revenue",
                MR["annualRevenue"], '"$"#,##0', "Belts, jerseys, books")
write_input_row(ws, 45, "International licensing",
                LC["annualRevenue"], '"$"#,##0', "Format · territories")
label_cell(ws, 46, "Off-platform replay %")
for c in (2, 3, 4):
    input_cell(ws, 46, c, AS_["offPlatformPctOfReplay"], "0%")
note(ws, 46, "FB · TikTok · YouTube share of replay")
label_cell(ws, 47, "Ad-share CPM (off-platform)")
for c in (2, 3, 4):
    input_cell(ws, 47, c, AS_["cpm"], '"$"#,##0.00')
note(ws, 47, "YouTube AdSense estimate")


# ── Tab 3: Bouts & Audience ──────────────────────────────────────────────
ws = wb.create_sheet("Bouts & Audience")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 32
for col in "BCDE":
    ws.column_dimensions[col].width = 20

kicker(ws, 1, "VOLUME")
title(ws, 2, "Bouts and viewers, by year")
header_row(ws, 4, ["Metric", "2026", "2027", "2028", "3-Yr Total"])

label_cell(ws, 6, "Bouts per year")
calc_cell(ws, 6, 2, "=Assumptions!B7", "#,##0")
calc_cell(ws, 6, 3, "=Assumptions!C7", "#,##0")
calc_cell(ws, 6, 4, "=Assumptions!D7", "#,##0")
calc_cell(ws, 6, 5, "=SUM(B6:D6)", "#,##0", total=True)

label_cell(ws, 7, "Avg concurrent live viewers / bout")
calc_cell(ws, 7, 2, "=Assumptions!B8", "#,##0")
calc_cell(ws, 7, 3, "=Assumptions!C8", "#,##0")
calc_cell(ws, 7, 4, "=Assumptions!D8", "#,##0")

label_cell(ws, 8, "Annual viewer-bouts (live)")
calc_cell(ws, 8, 2, "=B6*B7", "#,##0")
calc_cell(ws, 8, 3, "=C6*C7", "#,##0")
calc_cell(ws, 8, 4, "=D6*D7", "#,##0")
calc_cell(ws, 8, 5, "=SUM(B8:D8)", "#,##0", total=True)

label_cell(ws, 9, "Voter conversion rate")
calc_cell(ws, 9, 2, "=Assumptions!B9", "0.0%")
calc_cell(ws, 9, 3, "=Assumptions!C9", "0.0%")
calc_cell(ws, 9, 4, "=Assumptions!D9", "0.0%")

label_cell(ws, 10, "Votes per bout")
calc_cell(ws, 10, 2, "=B7*B9", "#,##0")
calc_cell(ws, 10, 3, "=C7*C9", "#,##0")
calc_cell(ws, 10, 4, "=D7*D9", "#,##0")

label_cell(ws, 11, "Annual votes", total=True)
calc_cell(ws, 11, 2, "=B6*B10", "#,##0", total=True)
calc_cell(ws, 11, 3, "=C6*C10", "#,##0", total=True)
calc_cell(ws, 11, 4, "=D6*D10", "#,##0", total=True)
calc_cell(ws, 11, 5, "=SUM(B11:D11)", "#,##0", total=True)

label_cell(ws, 13, "Avg minutes watched / live viewer")
calc_cell(ws, 13, 2, "=Assumptions!B34", "#,##0")
calc_cell(ws, 13, 3, "=Assumptions!C34", "#,##0")
calc_cell(ws, 13, 4, "=Assumptions!D34", "#,##0")

# Mux annual viewer-hours = bouts × (live viewers × live-min + live × replay-mult × replay-min × on-platform-pct) / 60
# Where on-platform-pct = (1 - off-platform-pct) on Assumptions row 46.
label_cell(ws, 14, "Annual viewer-hours delivered (Mux)", total=True)
for c, col in [(2, "B"), (3, "C"), (4, "D")]:
    formula = (
        f"=B6*({col}7*{col}13+{col}7*Assumptions!{col}35*Assumptions!{col}36*"
        f"(1-Assumptions!{col}46))/60"
    )
    if col == "B":
        formula = f"=B6*(B7*B13+B7*Assumptions!B35*Assumptions!B36*(1-Assumptions!B46))/60"
    elif col == "C":
        formula = f"=C6*(C7*C13+C7*Assumptions!C35*Assumptions!C36*(1-Assumptions!C46))/60"
    else:
        formula = f"=D6*(D7*D13+D7*Assumptions!D35*Assumptions!D36*(1-Assumptions!D46))/60"
    calc_cell(ws, 14, c, formula, "#,##0", total=True)
calc_cell(ws, 14, 5, "=SUM(B14:D14)", "#,##0", total=True)

section_header(ws, 16, "REACH BEYOND LIVE", color=BLUE)

label_cell(ws, 17, "Post-live replay multiplier")
calc_cell(ws, 17, 2, "=Assumptions!B35", '#,##0"x"')
calc_cell(ws, 17, 3, "=Assumptions!C35", '#,##0"x"')
calc_cell(ws, 17, 4, "=Assumptions!D35", '#,##0"x"')

label_cell(ws, 18, "Replay viewers / bout")
calc_cell(ws, 18, 2, "=B7*B17", "#,##0")
calc_cell(ws, 18, 3, "=C7*C17", "#,##0")
calc_cell(ws, 18, 4, "=D7*D17", "#,##0")

label_cell(ws, 19, "Total reach / bout (live + replay)")
calc_cell(ws, 19, 2, "=B7+B18", "#,##0")
calc_cell(ws, 19, 3, "=C7+C18", "#,##0")
calc_cell(ws, 19, 4, "=D7+D18", "#,##0")

label_cell(ws, 20, "Annual total impressions", total=True)
calc_cell(ws, 20, 2, "=B6*B19", "#,##0", total=True)
calc_cell(ws, 20, 3, "=C6*C19", "#,##0", total=True)
calc_cell(ws, 20, 4, "=D6*D19", "#,##0", total=True)
calc_cell(ws, 20, 5, "=SUM(B20:D20)", "#,##0", total=True)


# ── Tab 4: P&L (3-Year) ──────────────────────────────────────────────────
ws = wb.create_sheet("P&L (3-Year)")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
for col in "BCDE":
    ws.column_dimensions[col].width = 18

kicker(ws, 1, "FULL P&L  ·  REVENUE  ·  COGS  ·  GROSS MARGIN  ·  OPEX  ·  EBITDA")
title(ws, 2, "P&L — 3-Year View")
header_row(ws, 4, ["", "2026", "2027", "2028", "3-Yr Total"])

# REVENUE — 7 lines + total
section_header(ws, 6, "REVENUE", color=RED)

# Row 7: Voting
label_cell(ws, 7, "Voting revenue")
calc_cell(ws, 7, 2, "='Bouts & Audience'!B11*Assumptions!B10", USD)
calc_cell(ws, 7, 3, "='Bouts & Audience'!C11*Assumptions!C10", USD)
calc_cell(ws, 7, 4, "='Bouts & Audience'!D11*Assumptions!D10", USD)
calc_cell(ws, 7, 5, "=SUM(B7:D7)", USD)

# Row 8: Sponsor
label_cell(ws, 8, "Sponsor revenue")
calc_cell(ws, 8, 2, "=Assumptions!B7*Assumptions!B14*Assumptions!B13", USD)
calc_cell(ws, 8, 3, "=Assumptions!C7*Assumptions!C14*Assumptions!C13", USD)
calc_cell(ws, 8, 4, "=Assumptions!D7*Assumptions!D14*Assumptions!D13", USD)
calc_cell(ws, 8, 5, "=SUM(B8:D8)", USD)

# Row 9: Premium subscriptions = subscribers × monthly price × 12
label_cell(ws, 9, "Premium subscriptions")
calc_cell(ws, 9, 2, "=Assumptions!B39*Assumptions!B40*12", USD)
calc_cell(ws, 9, 3, "=Assumptions!C39*Assumptions!C40*12", USD)
calc_cell(ws, 9, 4, "=Assumptions!D39*Assumptions!D40*12", USD)
calc_cell(ws, 9, 5, "=SUM(B9:D9)", USD)

# Row 10: Ticketing = events × seats × price
label_cell(ws, 10, "Live championship ticketing")
calc_cell(ws, 10, 2, "=Assumptions!B41*Assumptions!B42*Assumptions!B43", USD)
calc_cell(ws, 10, 3, "=Assumptions!C41*Assumptions!C42*Assumptions!C43", USD)
calc_cell(ws, 10, 4, "=Assumptions!D41*Assumptions!D42*Assumptions!D43", USD)
calc_cell(ws, 10, 5, "=SUM(B10:D10)", USD)

# Row 11: Merch
label_cell(ws, 11, "Merch")
calc_cell(ws, 11, 2, "=Assumptions!B44", USD)
calc_cell(ws, 11, 3, "=Assumptions!C44", USD)
calc_cell(ws, 11, 4, "=Assumptions!D44", USD)
calc_cell(ws, 11, 5, "=SUM(B11:D11)", USD)

# Row 12: Licensing
label_cell(ws, 12, "International licensing")
calc_cell(ws, 12, 2, "=Assumptions!B45", USD)
calc_cell(ws, 12, 3, "=Assumptions!C45", USD)
calc_cell(ws, 12, 4, "=Assumptions!D45", USD)
calc_cell(ws, 12, 5, "=SUM(B12:D12)", USD)

# Row 13: YouTube ad share = bouts × live × replay-mult × off-platform-pct × CPM / 1000
label_cell(ws, 13, "YouTube ad share")
calc_cell(ws, 13, 2,
          "=Assumptions!B7*Assumptions!B8*Assumptions!B35*Assumptions!B46*Assumptions!B47/1000", USD)
calc_cell(ws, 13, 3,
          "=Assumptions!C7*Assumptions!C8*Assumptions!C35*Assumptions!C46*Assumptions!C47/1000", USD)
calc_cell(ws, 13, 4,
          "=Assumptions!D7*Assumptions!D8*Assumptions!D35*Assumptions!D46*Assumptions!D47/1000", USD)
calc_cell(ws, 13, 5, "=SUM(B13:D13)", USD)

# Row 14: Total revenue
label_cell(ws, 14, "Total revenue", total=True)
calc_cell(ws, 14, 2, "=SUM(B7:B13)", USD, total=True)
calc_cell(ws, 14, 3, "=SUM(C7:C13)", USD, total=True)
calc_cell(ws, 14, 4, "=SUM(D7:D13)", USD, total=True)
calc_cell(ws, 14, 5, "=SUM(B14:D14)", USD, total=True)

# VARIABLE COSTS
section_header(ws, 16, "VARIABLE COSTS  ·  COGS", color=BLUE)

label_cell(ws, 17, "Mux delivery (live + on-platform replay)")
calc_cell(ws, 17, 2, "='Bouts & Audience'!B14*Assumptions!B17", USD)
calc_cell(ws, 17, 3, "='Bouts & Audience'!C14*Assumptions!C17", USD)
calc_cell(ws, 17, 4, "='Bouts & Audience'!D14*Assumptions!D17", USD)

label_cell(ws, 18, "Mux ingest")
calc_cell(ws, 18, 2, "=Assumptions!B7*Assumptions!B18", USD)
calc_cell(ws, 18, 3, "=Assumptions!C7*Assumptions!C18", USD)
calc_cell(ws, 18, 4, "=Assumptions!D7*Assumptions!D18", USD)

label_cell(ws, 19, "Stripe processing")
calc_cell(ws, 19, 2, "=B7*Assumptions!B19+'Bouts & Audience'!B11*Assumptions!B20", USD)
calc_cell(ws, 19, 3, "=C7*Assumptions!C19+'Bouts & Audience'!C11*Assumptions!C20", USD)
calc_cell(ws, 19, 4, "=D7*Assumptions!D19+'Bouts & Audience'!D11*Assumptions!D20", USD)

label_cell(ws, 20, "Charity payout (18% of vote rev)")
calc_cell(ws, 20, 2, "=B7*Assumptions!B21", USD)
calc_cell(ws, 20, 3, "=C7*Assumptions!C21", USD)
calc_cell(ws, 20, 4, "=D7*Assumptions!D21", USD)

label_cell(ws, 21, "Debater honorariums")
calc_cell(ws, 21, 2, "=Assumptions!B7*Assumptions!B22", USD)
calc_cell(ws, 21, 3, "=Assumptions!C7*Assumptions!C22", USD)
calc_cell(ws, 21, 4, "=Assumptions!D7*Assumptions!D22", USD)

label_cell(ws, 22, "Production")
calc_cell(ws, 22, 2, "=Assumptions!B7*Assumptions!B23", USD)
calc_cell(ws, 22, 3, "=Assumptions!C7*Assumptions!C23", USD)
calc_cell(ws, 22, 4, "=Assumptions!D7*Assumptions!D23", USD)

label_cell(ws, 23, "Total variable cost", total=True)
calc_cell(ws, 23, 2, "=SUM(B17:B22)", USD, total=True)
calc_cell(ws, 23, 3, "=SUM(C17:C22)", USD, total=True)
calc_cell(ws, 23, 4, "=SUM(D17:D22)", USD, total=True)
calc_cell(ws, 23, 5, "=SUM(B23:D23)", USD, total=True)

# GROSS MARGIN
section_header(ws, 25, "GROSS MARGIN", color=INK)

label_cell(ws, 26, "Gross profit", total=True)
calc_cell(ws, 26, 2, "=B14-B23", USD, total=True)
calc_cell(ws, 26, 3, "=C14-C23", USD, total=True)
calc_cell(ws, 26, 4, "=D14-D23", USD, total=True)
calc_cell(ws, 26, 5, "=SUM(B26:D26)", USD, total=True)

label_cell(ws, 27, "Gross margin %")
calc_cell(ws, 27, 2, "=IFERROR(B26/B14,0)", PCT)
calc_cell(ws, 27, 3, "=IFERROR(C26/C14,0)", PCT)
calc_cell(ws, 27, 4, "=IFERROR(D26/D14,0)", PCT)

# FIXED OPEX
section_header(ws, 29, "FIXED OPEX", color=BLUE)

label_cell(ws, 30, "Headcount cost")
calc_cell(ws, 30, 2, "=Assumptions!B26*Assumptions!B27", USD)
calc_cell(ws, 30, 3, "=Assumptions!C26*Assumptions!C27", USD)
calc_cell(ws, 30, 4, "=Assumptions!D26*Assumptions!D27", USD)

label_cell(ws, 31, "Paid search & acquisition")
calc_cell(ws, 31, 2, "=Assumptions!B28", USD)
calc_cell(ws, 31, 3, "=Assumptions!C28", USD)
calc_cell(ws, 31, 4, "=Assumptions!D28", USD)

label_cell(ws, 32, "Brand & creator marketing")
calc_cell(ws, 32, 2, "=Assumptions!B29", USD)
calc_cell(ws, 32, 3, "=Assumptions!C29", USD)
calc_cell(ws, 32, 4, "=Assumptions!D29", USD)

label_cell(ws, 33, "G&A, tooling, infra")
calc_cell(ws, 33, 2, "=Assumptions!B30", USD)
calc_cell(ws, 33, 3, "=Assumptions!C30", USD)
calc_cell(ws, 33, 4, "=Assumptions!D30", USD)

label_cell(ws, 34, "Legal, accounting, insurance")
calc_cell(ws, 34, 2, "=Assumptions!B31", USD)
calc_cell(ws, 34, 3, "=Assumptions!C31", USD)
calc_cell(ws, 34, 4, "=Assumptions!D31", USD)

label_cell(ws, 35, "Total fixed opex", total=True)
calc_cell(ws, 35, 2, "=SUM(B30:B34)", USD, total=True)
calc_cell(ws, 35, 3, "=SUM(C30:C34)", USD, total=True)
calc_cell(ws, 35, 4, "=SUM(D30:D34)", USD, total=True)
calc_cell(ws, 35, 5, "=SUM(B35:D35)", USD, total=True)

# EBITDA
section_header(ws, 37, "EBITDA", color=INK)

label_cell(ws, 38, "EBITDA", total=True)
calc_cell(ws, 38, 2, "=B26-B35", USD, total=True)
calc_cell(ws, 38, 3, "=C26-C35", USD, total=True)
calc_cell(ws, 38, 4, "=D26-D35", USD, total=True)
calc_cell(ws, 38, 5, "=SUM(B38:D38)", USD, total=True)

label_cell(ws, 39, "EBITDA margin %")
calc_cell(ws, 39, 2, "=IFERROR(B38/B14,0)", PCT)
calc_cell(ws, 39, 3, "=IFERROR(C38/C14,0)", PCT)
calc_cell(ws, 39, 4, "=IFERROR(D38/D14,0)", PCT)

ws.cell(row=42, column=1,
        value="Pre-tax. Excludes working-capital float and tax.").font = NOTE


# ── Tab 5: Unit Economics ────────────────────────────────────────────────
ws = wb.create_sheet("Unit Economics")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
for col in "BCD":
    ws.column_dimensions[col].width = 18

kicker(ws, 1, "PER-BOUT CONTRIBUTION MARGIN AT THREE AUDIENCE SCALES")
title(ws, 2, "Unit Economics — per bout")
header_row(ws, 4, ["", "Year 1 scale", "Year 2 scale", "Year 3 scale"], color=INK)

label_cell(ws, 6, "Avg viewers / bout")
calc_cell(ws, 6, 2, "=Assumptions!B8", "#,##0")
calc_cell(ws, 6, 3, "=Assumptions!C8", "#,##0")
calc_cell(ws, 6, 4, "=Assumptions!D8", "#,##0")

label_cell(ws, 7, "Voters / bout")
calc_cell(ws, 7, 2, "=B6*Assumptions!B9", "#,##0")
calc_cell(ws, 7, 3, "=C6*Assumptions!C9", "#,##0")
calc_cell(ws, 7, 4, "=D6*Assumptions!D9", "#,##0")

# Revenue / bout (excludes one-off lines like ticketing/merch/licensing)
section_header(ws, 9, "REVENUE / BOUT", color=RED, span_to=4)

label_cell(ws, 10, "Voting / bout")
calc_cell(ws, 10, 2, "=B7*Assumptions!B10", USD)
calc_cell(ws, 10, 3, "=C7*Assumptions!C10", USD)
calc_cell(ws, 10, 4, "=D7*Assumptions!D10", USD)

label_cell(ws, 11, "Sponsor / bout (allocated)")
calc_cell(ws, 11, 2, "=Assumptions!B13*Assumptions!B14", USD)
calc_cell(ws, 11, 3, "=Assumptions!C13*Assumptions!C14", USD)
calc_cell(ws, 11, 4, "=Assumptions!D13*Assumptions!D14", USD)

label_cell(ws, 12, "Ad share / bout")
calc_cell(ws, 12, 2, "=Assumptions!B8*Assumptions!B35*Assumptions!B46*Assumptions!B47/1000", USD)
calc_cell(ws, 12, 3, "=Assumptions!C8*Assumptions!C35*Assumptions!C46*Assumptions!C47/1000", USD)
calc_cell(ws, 12, 4, "=Assumptions!D8*Assumptions!D35*Assumptions!D46*Assumptions!D47/1000", USD)

label_cell(ws, 13, "Total revenue / bout", total=True)
calc_cell(ws, 13, 2, "=SUM(B10:B12)", USD, total=True)
calc_cell(ws, 13, 3, "=SUM(C10:C12)", USD, total=True)
calc_cell(ws, 13, 4, "=SUM(D10:D12)", USD, total=True)

# Variable cost / bout
section_header(ws, 15, "VARIABLE COST / BOUT", color=BLUE, span_to=4)

label_cell(ws, 16, "Mux delivery (live + on-platform replay)")
calc_cell(ws, 16, 2,
          "=B6*(Assumptions!B34+Assumptions!B35*Assumptions!B36*(1-Assumptions!B46))/60*Assumptions!B17", USD)
calc_cell(ws, 16, 3,
          "=C6*(Assumptions!C34+Assumptions!C35*Assumptions!C36*(1-Assumptions!C46))/60*Assumptions!C17", USD)
calc_cell(ws, 16, 4,
          "=D6*(Assumptions!D34+Assumptions!D35*Assumptions!D36*(1-Assumptions!D46))/60*Assumptions!D17", USD)

label_cell(ws, 17, "Mux ingest")
calc_cell(ws, 17, 2, "=Assumptions!B18", USD)
calc_cell(ws, 17, 3, "=Assumptions!C18", USD)
calc_cell(ws, 17, 4, "=Assumptions!D18", USD)

label_cell(ws, 18, "Stripe processing")
calc_cell(ws, 18, 2, "=B10*Assumptions!B19+B7*Assumptions!B20", USD)
calc_cell(ws, 18, 3, "=C10*Assumptions!C19+C7*Assumptions!C20", USD)
calc_cell(ws, 18, 4, "=D10*Assumptions!D19+D7*Assumptions!D20", USD)

label_cell(ws, 19, "Charity payout")
calc_cell(ws, 19, 2, "=B10*Assumptions!B21", USD)
calc_cell(ws, 19, 3, "=C10*Assumptions!C21", USD)
calc_cell(ws, 19, 4, "=D10*Assumptions!D21", USD)

label_cell(ws, 20, "Debater honorariums")
calc_cell(ws, 20, 2, "=Assumptions!B22", USD)
calc_cell(ws, 20, 3, "=Assumptions!C22", USD)
calc_cell(ws, 20, 4, "=Assumptions!D22", USD)

label_cell(ws, 21, "Production")
calc_cell(ws, 21, 2, "=Assumptions!B23", USD)
calc_cell(ws, 21, 3, "=Assumptions!C23", USD)
calc_cell(ws, 21, 4, "=Assumptions!D23", USD)

label_cell(ws, 22, "Total variable cost / bout", total=True)
calc_cell(ws, 22, 2, "=SUM(B16:B21)", USD, total=True)
calc_cell(ws, 22, 3, "=SUM(C16:C21)", USD, total=True)
calc_cell(ws, 22, 4, "=SUM(D16:D21)", USD, total=True)

# Contribution
section_header(ws, 24, "CONTRIBUTION MARGIN / BOUT", color=INK, span_to=4)

label_cell(ws, 25, "Contribution $", total=True)
calc_cell(ws, 25, 2, "=B13-B22", USD, total=True)
calc_cell(ws, 25, 3, "=C13-C22", USD, total=True)
calc_cell(ws, 25, 4, "=D13-D22", USD, total=True)

label_cell(ws, 26, "Contribution %")
calc_cell(ws, 26, 2, "=IFERROR(B25/B13,0)", PCT)
calc_cell(ws, 26, 3, "=IFERROR(C25/C13,0)", PCT)
calc_cell(ws, 26, 4, "=IFERROR(D25/D13,0)", PCT)

ws.cell(row=28, column=1,
        value="Excludes one-off revenue lines (premium subs, ticketing, merch, licensing) and fixed opex.").font = NOTE


# ── Save ─────────────────────────────────────────────────────────────────
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH.relative_to(APP_ROOT.parent)} ({SCENARIO_LABEL} scenario)")
